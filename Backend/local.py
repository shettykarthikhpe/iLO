# import csv
import pandas as pd
import redfish
from pydantic import BaseModel
import redfish
import openpyxl
import os
import csv
import pandas as pd

# # read specific columns of csv file using Pandas
# filename = "../csv/customers-100.csv"
# df = pd.read_csv(filename)
# data =[] 
# for rows in df:
#     data.append(str(rows))


# datas = [
#     {1, 'hgwbnw', "karthik", "hello", "hpe", "bangalore", "india", "456789", "993848393", "kskk@gmail.com", "12-122-2222", "hello"}
# ]

# # writing to csv file
# with open(filename, 'a') as csvfile:
#     # creating a csv writer object
#     csvwriter = csv.writer(csvfile)

#     # # writing the fields
#     # csvwriter.writerow(datas)

#     # writing the data rows
#     csvwriter.writerows(datas)

import openpyxl
import csv
import os

def file_converter(file_name):
	file_extension = os.path.splitext(file_name)[1]
	file_original_name = os.path.splitext(file_name)[0]
	if(file_extension == '.xlsx'):
		# Load the Excel file
		wb = openpyxl.load_workbook(f'../uploads/{file_original_name}.xlsx')
		sheet = wb.active

		# Open a CSV file to write
		with open(f'../csv/{file_original_name}.csv', 'w', newline="") as f:
			c = csv.writer(f)

			# Write the rows from the Excel sheet to the CSV file
			for row in sheet.iter_rows(values_only=True):
				c.writerow(row)
	else:
		print("proper csv")

# # function to geth the drive list
# Drives= []
# def DriveGetter():
#     url = "10.132.147.215"
#     username = "Administrator"
#     password = "GXJYN722"
#     # Create a Redfish client object and connect to the API
#     client = redfish.redfish_client(base_url=url, username=username, password=password)
#     try:
#         client.login(auth=redfish.AuthMethod.SESSION)
#         # Perform some tasks using the Redfish API
#         response = client.get('/redfish/v1/Systems/1/Storage/')
#         if response.status == 200:
#             for murl in response.dict["Members"]:
#                 resp = client.get(murl["@odata.id"])
#                 for mul in resp.dict["Drives"]:
#                     re= client.get(mul["@odata.id"])
#                     Drives.append(re.dict['SerialNumber'])
#             return ([Drives])
#         else:
#             return ([])
#     except Exception as e:
#          return ([])


# filename = "../uploads/iLO_OC_Accessories_Inventory_sheet.csv"
# df = pd.read_csv(filename, encoding="ISO-8859-1")

# # ActualDrives= DriveGetter()
# drive_str_no = df[df['Type'].str.lower()=='memory' ]['Part number'] == 'P43328-B21'
# # print(drive_str_no.isin(ActualDrives))
# # print(drive_str_no)
# # matchedDrives= []
# # matched = drive_str_no[drive_str_no.isin(ActualDrives[0])]
# print("macthed are",drive_str_no)

# for rows in df:
#     if rows == 'Type':
#         print(df['Type'])
#         for item in df[rows]:
#             if item == 'Drive':
#         					print(item)
          
          
#         if df[rows][1] == 'Drive':  
#     if df[rows]== 'Drive':
        
# print(df['Type']== 'Drive')
# # iterating the rows
# for rows in df:
#     if rows == 'SR no':
#         if df['Type'] == "Drive":
# 			for item in df[rows]:
# 				print(item)
#         	if item in ActualDrives:
#             		print(df[rows])





from pydantic import BaseModel
from fastapi import FastAPI, HTTPException, File, UploadFile

app = FastAPI()
 
class IP(BaseModel):
    partNumber: str

from pymongo import MongoClient

# Create a connection to the MongoDB server
client = MongoClient("mongodb+srv://abhishekdrai85:Abhishek29@cluster0.4kjtl.mongodb.net/hp?retryWrites=true&w=majority&appName=Cluster0")

# Access a specific database
db = client["hp"]

# Access a specific collection within the database
collection = db["suts"]

processorSet= []
result =[]
@app.post('/local')
def Local(data:IP):
    # print(data)
    partNumber = data.partNumber
    ipLists = collection.find_one({'userId':"85oiv2tqz4"})
    for listItem in ipLists['sut']:
        client = redfish.redfish_client(base_url=listItem['ip'], username=listItem['username'], password=listItem['password'])
        # Attempt to login with a timeout of 10 seconds
        try:
            client.login(auth=redfish.AuthMethod.SESSION)
            # Perform some tasks using the Redfish API
            # response = client.get('/redfish/v1/Systems/1/Memory/proc1dimm1')
            response = client.get("/redfish/v1/Systems/1/Memory")
            
            if response.status == 200:
                processorSet.append(response.dict['Members'])
                for item in processorSet[0]:
                    resp = client.get(item['@odata.id'])
                    print(resp.dict['Oem']['Hpe']['PartNumber'])
                    if (str(partNumber) == resp.dict['Oem']['Hpe']['PartNumber']):
                        result.append(listItem['ip'])
        except Exception as e:
            return ({ "error": e })
    return (result)

class IP(BaseModel):
    partNumber: str
    filename:str

class File(BaseModel):
    filename:str

def file_converter(file_name):
	file_extension = os.path.splitext(file_name)[1]
	file_original_name = os.path.splitext(file_name)[0]
	if(file_extension == '.xlsx'):
		# Load the Excel file
		wb = openpyxl.load_workbook(f'../uploads/{file_original_name}.xlsx')
		sheet = wb.active

		# Open a CSV file to write
		with open(f'../uploads/{file_original_name}.csv', 'w', newline="") as f:
			c = csv.writer(f)

			# Write the rows from the Excel sheet to the CSV file
			for row in sheet.iter_rows(values_only=True):
				c.writerow(row)
	else:
		print("proper csv")

@app.post("/upload")
def converter(data:File):
    file_name = data.filename
    file_converter(file_name)
    print(file_name)

@app.post('/localIn')
def LocalIn(data:IP):
    filename = f"../uploads/{data.filename}.csv"
    df = pd.read_csv(filename, encoding="ISO-8859-1")
    drive_str_no = df[df['Type'].str.lower()=='memory' ]['Part number'] == data.partNumber
    for item in drive_str_no:
        if item == True:
            return (True)
    else:
        return (False)
    
    